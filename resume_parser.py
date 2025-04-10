import os
import spacy
import pandas as pd
from pathlib import Path
from docx import Document
import openai  # Import OpenAI library
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import json  # Import the JSON library for parsing the API key file

# Load the NLP model (use 'en_core_web_sm' for simplicity)
nlp = spacy.load("en_core_web_sm")

def load_api_key(file_path):
    """
    Load the OpenAI API key from a JSON file. Extract the value of the 'openai' key.
    """
    try:
        with open(file_path, "r") as file:
            data = json.load(file)  # Parse the JSON file
            print (data.get("auto_project"))
            return data.get("auto_project")  # Retrieve the 'openai' key
    except Exception as e:
        print(f"Error loading API key from {file_path}: {e}")
        return None

# Load the OpenAI API key from the specified file
api_key_file = "/Users/sarenah/Documents/ST/MyProject/API_KEY.txt"
openai.api_key = load_api_key(api_key_file)

if not openai.api_key:
    raise ValueError("OpenAI API key could not be loaded. Please check the file path and contents.")

def extract_resume_data_with_gpt(text):
    """
    Use OpenAI GPT to extract key information from resume text.
    """
    prompt = f"""
    Extract the following information from the resume text:
    - Name
    - Email
    - Phone
    - Location
    - Education
    - Experience
    - Skills
    - Autobiography

    Resume Text:
    {text}

    Provide the information in JSON format with keys: Name, Email, Phone, Location, Education, Experience, Skills, Autobiography.
    """
    try:
        response = openai.ChatCompletion.create(
            model="gpt-4",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=1000,
            temperature=0
        )
        result = response['choices'][0]['message']['content']
        return json.loads(result)  # Safely parse JSON string to dictionary
    except json.JSONDecodeError as e:
        print(f"Error parsing JSON response: {e}")
        return {
            "Name": None,
            "Email": None,
            "Phone": None,
            "Location": None,
            "Education": None,
            "Experience": None,
            "Skills": None,
            "Autobiography": None
        }
    except Exception as e:
        print(f"Error with OpenAI API: {e}")
        return {
            "Name": None,
            "Email": None,
            "Phone": None,
            "Location": None,
            "Education": None,
            "Experience": None,
            "Skills": None,
            "Autobiography": None
        }

def analyze_personality_with_gpt(autobiography):
    """
    Use OpenAI GPT to analyze personality traits from the autobiography.
    """
    prompt = f"""
    Analyze the personality traits of the individual based on the following autobiography:
    {autobiography}

    Provide a concise summary of the personality traits in one sentence.
    """
    try:
        response = openai.ChatCompletion.create(
            model="gpt-4",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=150,
            temperature=0
        )
        result = response['choices'][0]['message']['content']
        return result.strip()  # Return the personality analysis
    except Exception as e:
        print(f"Error with OpenAI API for personality analysis: {e}")
        return "Personality analysis unavailable"

def read_docx(file_path):
    """
    Read text from a .docx file.
    """
    doc = Document(file_path)
    return "\n".join([paragraph.text for paragraph in doc.paragraphs])

def parse_user_demands_with_gpt(user_input):
    """
    Use OpenAI GPT to parse natural language user input into structured demands.
    """
    prompt = f"""
    Parse the following user input into structured demands for a candidate:
    "{user_input}"

    Extract the demands into categories: Skills, Location, Experience, and Other.
    Provide the result in JSON format with keys: Skills, Location, Experience, Other.
    """
    try:
        response = openai.ChatCompletion.create(
            model="gpt-4",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=200,
            temperature=0
        )
        result = response['choices'][0]['message']['content']
        return json.loads(result)  # Safely parse JSON string to dictionary
    except json.JSONDecodeError as e:
        print(f"Error parsing JSON response: {e}")
        return {
            "Skills": [],
            "Location": [],
            "Experience": [],
            "Other": []
        }
    except Exception as e:
        print(f"Error with OpenAI API for parsing demands: {e}")
        return {
            "Skills": [],
            "Location": [],
            "Experience": [],
            "Other": []
        }

def get_user_demands():
    """
    Allow the user to input demands for the type of person they need in natural language.
    """
    print("Enter your demands for the candidate (e.g., 'I need someone skilled in Python, based in Dublin, with 5 years of experience.'):")
    user_input = input("Your demands: ").strip()
    return parse_user_demands_with_gpt(user_input)

def is_candidate_qualified_with_gpt(resume_data, demands):
    """
    Use OpenAI GPT to check if a candidate meets the user's demands with semantic understanding.
    Also, return the matched keywords based on semantic similarity.
    """
    prompt = f"""
    Evaluate if the following candidate's information matches the user's demands:
    
    User Demands:
    Skills: {', '.join(demands.get("Skills", []))}
    Location: {', '.join(demands.get("Location", []))}
    Experience: {', '.join(demands.get("Experience", []))}
    Other: {', '.join(demands.get("Other", []))}

    Candidate Information:
    Skills: {resume_data.get("Skills", "")}
    Location: {resume_data.get("Location", "")}
    Experience: {resume_data.get("Experience", "")}
    Education: {resume_data.get("Education", "")}
    Autobiography: {resume_data.get("Autobiography", "")}

    Respond with "Yes" if the candidate matches the user's demands based on semantic understanding, even if the keywords are not exactly the same.
    Consider related terms, synonyms, and broader meanings.
    Also, provide a list of keywords or phrases from the candidate's information that semantically match the user's demands.
    Format your response as:
    "Yes: [matched_keyword1, matched_keyword2, ...]" or "No: []".
    """
    try:
        response = openai.ChatCompletion.create(
            model="gpt-4",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=150,
            temperature=0
        )
        result = response['choices'][0]['message']['content'].strip()
        if result.lower().startswith("yes:"):
            matched_keywords = result[4:].strip()  # Extract the matched keywords
            return True, matched_keywords
        return False, "[]"
    except Exception as e:
        print(f"Error with OpenAI API for candidate qualification: {e}")
        return False, "[]"

def highlight_qualified_candidates(output_file, qualified_rows):
    """
    Highlight qualified candidates in the Excel file.
    """
    wb = load_workbook(output_file)
    ws = wb.active
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow highlight

    for row in qualified_rows:
        for cell in ws[row]:
            cell.fill = fill

    wb.save(output_file)
    print("Qualified candidates highlighted in the Excel file.")

def process_resumes(input_dir, output_file):
    """
    Process all resumes in the input directory and save the data to an Excel file.
    """
    demands = get_user_demands()  # Get user demands
    data = []
    qualified_rows = []

    for idx, file in enumerate(Path(input_dir).glob("*.docx"), start=2):  # Process .docx files, start=2 for Excel rows
        text = read_docx(file)  # Read text from .docx
        resume_data = extract_resume_data_with_gpt(text)  # Use GPT for extraction
        resume_data["File"] = file.name

        # Analyze personality from autobiography
        autobiography = resume_data.get("Autobiography", "")
        resume_data["Personality"] = analyze_personality_with_gpt(autobiography)

        # Check if the candidate meets the demands using GPT
        is_qualified, matched_keywords = is_candidate_qualified_with_gpt(resume_data, demands)
        resume_data["Matched Keywords"] = matched_keywords  # Add matched keywords to the data
        if is_qualified:
            qualified_rows.append(idx)  # Track row number for highlighting

        data.append(resume_data)

    # Save to Excel
    df = pd.DataFrame(data)
    df.to_excel(output_file, index=False)
    print(f"Data saved to {output_file}")

    # Highlight qualified candidates
    highlight_qualified_candidates(output_file, qualified_rows)

if __name__ == "__main__":
    input_directory = "/Users/sarenah/Documents/ST/愛爾蘭/Intelligent Agents and Process Automation/TEMP"
    output_excel = "/Users/sarenah/Documents/ST/愛爾蘭/Intelligent Agents and Process Automation/TEMP/resumes.xlsx"
    process_resumes(input_directory, output_excel)
